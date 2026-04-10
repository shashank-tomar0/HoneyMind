"""
Canary Token Routes
===================
API endpoints for the canary token system:
    - Generate bait files with embedded tracking beacons
    - Receive canary ping callbacks (beacon hits)
    - List all canary hits
"""

from __future__ import annotations

import logging
import uuid
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path

from flask import Blueprint, request, jsonify, send_file, current_app

from honeyshield.backend.models import db, CanaryHit, AttackSession

logger = logging.getLogger(__name__)

canary_bp = Blueprint("canary", __name__, url_prefix="/api/canary")

# 1x1 transparent GIF for beacon response
TRANSPARENT_GIF = (
    b"\x47\x49\x46\x38\x39\x61\x01\x00\x01\x00\x80\x00\x00"
    b"\xff\xff\xff\x00\x00\x00\x21\xf9\x04\x00\x00\x00\x00"
    b"\x00\x2c\x00\x00\x00\x00\x01\x00\x01\x00\x00\x02\x02"
    b"\x44\x01\x00\x3b"
)

# In-memory token registry: {token_id: metadata}
_token_registry: dict[str, dict] = {}


# ── Beacon Ping Endpoint ─────────────────────────────────────────────────


@canary_bp.route("/ping/<token_id>", methods=["GET"])
def canary_ping(token_id: str):
    """
    Canary token callback — triggered when attacker opens a bait file.
    Captures their real IP (potentially bypassing VPN) and logs the hit.

    Returns a 1x1 transparent GIF to remain invisible.
    """
    real_ip = request.headers.get("X-Forwarded-For", request.remote_addr)
    real_ua = request.headers.get("User-Agent", "")
    referrer = request.headers.get("Referer", "")

    # Look up token metadata
    token_meta = _token_registry.get(token_id, {})

    # Record the hit
    hit = CanaryHit(
        token_id=token_id,
        session_id=token_meta.get("session_id"),
        bait_file_type=token_meta.get("file_type", "unknown"),
        bait_file_name=token_meta.get("file_name", "unknown"),
        real_ip=real_ip,
        real_user_agent=real_ua,
        real_referrer=referrer,
    )
    db.session.add(hit)

    # If linked to a session, update the session with real IP
    if token_meta.get("session_id"):
        session = AttackSession.query.get(token_meta["session_id"])
        if session:
            session.real_ip_via_canary = real_ip
            logger.warning(
                "CANARY HIT: token=%s real_ip=%s (session=%s)",
                token_id, real_ip, session.id,
            )

    db.session.commit()

    logger.warning(
        "CANARY TOKEN TRIGGERED: token=%s real_ip=%s ua=%s",
        token_id, real_ip, real_ua[:80],
    )

    # Return invisible 1x1 GIF
    return send_file(
        BytesIO(TRANSPARENT_GIF),
        mimetype="image/gif",
        as_attachment=False,
    )


# ── Generate Bait File ──────────────────────────────────────────────────


@canary_bp.route("/generate", methods=["GET", "POST"])
def generate_bait():
    """
    Generate a canary-embedded bait file.

    Query params or JSON body:
        type: html|xlsx|pdf (default: html)
        session_id: optional, link to an attack session
        filename: optional, custom filename for the bait
    """
    if request.method == "POST":
        data = request.get_json(force=True, silent=True) or {}
    else:
        data = dict(request.args)

    file_type = data.get("type", "html").lower()
    session_id = data.get("session_id")
    custom_name = data.get("filename")

    if file_type not in ("html", "xlsx", "pdf", "docx"):
        return jsonify({"error": "Unsupported file type. Use: html, xlsx, pdf, docx"}), 400

    # Generate unique token
    token_id = str(uuid.uuid4())
    server_url = current_app.config.get("HONEYSHIELD_CONFIG", {}).get(
        "canary", {}
    ).get("server_url", "http://localhost:5000")
    beacon_url = f"{server_url}/api/canary/ping/{token_id}"

    # Determine filename
    filenames = {
        "html": custom_name or "employee_directory.html",
        "xlsx": custom_name or "db_dump_2024.xlsx",
        "pdf": custom_name or "network_topology.pdf",
        "docx": custom_name or "credentials.docx",
    }
    filename = filenames[file_type]

    # Register token
    _token_registry[token_id] = {
        "file_type": file_type,
        "file_name": filename,
        "session_id": session_id,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "beacon_url": beacon_url,
    }

    # Generate the bait file
    if file_type == "html":
        content = _generate_html_bait(token_id, beacon_url)
        mimetype = "text/html"
    elif file_type == "xlsx":
        content = _generate_xlsx_bait(token_id, beacon_url)
        mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    elif file_type == "pdf":
        content = _generate_pdf_bait(token_id, beacon_url)
        mimetype = "application/pdf"
    else:
        content = _generate_html_bait(token_id, beacon_url)
        mimetype = "text/html"

    logger.info("Generated bait file: type=%s token=%s name=%s",
                file_type, token_id, filename)

    return send_file(
        BytesIO(content),
        mimetype=mimetype,
        as_attachment=True,
        download_name=filename,
    )


@canary_bp.route("/generate/info", methods=["POST"])
def generate_bait_info():
    """
    Generate a canary token and return its metadata (without downloading).
    Useful for embedding into honeypot services programmatically.
    """
    data = request.get_json(force=True, silent=True) or {}
    file_type = data.get("type", "html")
    session_id = data.get("session_id")

    token_id = str(uuid.uuid4())
    server_url = current_app.config.get("HONEYSHIELD_CONFIG", {}).get(
        "canary", {}
    ).get("server_url", "http://localhost:5000")
    beacon_url = f"{server_url}/api/canary/ping/{token_id}"

    _token_registry[token_id] = {
        "file_type": file_type,
        "session_id": session_id,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "beacon_url": beacon_url,
    }

    return jsonify({
        "token_id": token_id,
        "beacon_url": beacon_url,
        "file_type": file_type,
        "session_id": session_id,
    }), 200


# ── List Canary Hits ─────────────────────────────────────────────────────


@canary_bp.route("/hits", methods=["GET"])
def list_canary_hits():
    """List all triggered canary token hits."""
    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 50, type=int)

    pagination = CanaryHit.query.order_by(
        CanaryHit.timestamp.desc()
    ).paginate(page=page, per_page=per_page, error_out=False)

    return jsonify({
        "hits": [h.to_dict() for h in pagination.items],
        "total": pagination.total,
        "page": pagination.page,
    }), 200


@canary_bp.route("/tokens", methods=["GET"])
def list_tokens():
    """List all generated canary tokens."""
    return jsonify({
        "tokens": _token_registry,
        "total": len(_token_registry),
    }), 200


# ── Bait File Generators ─────────────────────────────────────────────────


def _generate_html_bait(token_id: str, beacon_url: str) -> bytes:
    """Generate an HTML file with an embedded image beacon."""
    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <title>Internal Employee Directory — CONFIDENTIAL</title>
    <style>
        body {{ font-family: 'Segoe UI', Tahoma, sans-serif; margin: 40px;
               background: #f5f5f5; color: #333; }}
        h1 {{ color: #1a237e; border-bottom: 2px solid #1a237e; padding-bottom: 10px; }}
        table {{ border-collapse: collapse; width: 100%; margin-top: 20px;
                background: white; box-shadow: 0 1px 3px rgba(0,0,0,0.1); }}
        th {{ background: #1a237e; color: white; padding: 12px 15px; text-align: left; }}
        td {{ padding: 10px 15px; border-bottom: 1px solid #eee; }}
        tr:hover {{ background: #f0f0ff; }}
        .confidential {{ color: #d32f2f; font-weight: bold; font-size: 12px; }}
    </style>
</head>
<body>
    <p class="confidential">⚠ CONFIDENTIAL — INTERNAL USE ONLY</p>
    <h1>Employee Directory — 2024</h1>
    <table>
        <tr><th>Name</th><th>Department</th><th>Email</th><th>Phone</th><th>Access Level</th></tr>
        <tr><td>John Richardson</td><td>IT Admin</td><td>j.richardson@corp.internal</td><td>+1-555-0147</td><td>Root</td></tr>
        <tr><td>Sarah Chen</td><td>DevOps</td><td>s.chen@corp.internal</td><td>+1-555-0238</td><td>Sudo</td></tr>
        <tr><td>Mike Patel</td><td>Database</td><td>m.patel@corp.internal</td><td>+1-555-0329</td><td>DBA</td></tr>
        <tr><td>Lisa Anderson</td><td>Security</td><td>l.anderson@corp.internal</td><td>+1-555-0410</td><td>Admin</td></tr>
        <tr><td>David Kim</td><td>Network</td><td>d.kim@corp.internal</td><td>+1-555-0511</td><td>NetAdmin</td></tr>
    </table>
    <p style="margin-top: 40px; font-size: 11px; color: #999;">
        Last updated: 2024-03-15 | Generated by HR-Docs v3.2.1
    </p>
    <!-- Canary beacon -->
    <img src="{beacon_url}" width="1" height="1" style="position:absolute;left:-9999px;" alt="">
</body>
</html>"""
    return html.encode("utf-8")


def _generate_xlsx_bait(token_id: str, beacon_url: str) -> bytes:
    """
    Generate a minimal XLSX with an external data connection beacon.
    Falls back to HTML-in-XLSX if openpyxl not available.
    """
    try:
        import openpyxl
        from openpyxl.worksheet.hyperlink import Hyperlink

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "DB_Backup_2024"

        # Headers
        headers = ["ID", "Username", "Password_Hash", "Email", "Role", "Last_Login"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # Fake data rows
        fake_data = [
            [1, "admin", "5f4dcc3b5aa765d61d8327deb882cf99", "admin@corp.local", "superadmin", "2024-03-14"],
            [2, "dbadmin", "e99a18c428cb38d5f260853678922e03", "db@corp.local", "dba", "2024-03-13"],
            [3, "sysadmin", "d8578edf8458ce06fbc5bb76a58c5ca4", "sys@corp.local", "root", "2024-03-12"],
            [4, "backup_svc", "25d55ad283aa400af464c76d713c07ad", "backup@corp.local", "service", "2024-03-11"],
            [5, "deploy", "098f6bcd4621d373cade4e832627b4f6", "deploy@corp.local", "deploy", "2024-03-10"],
        ]
        for row_num, row_data in enumerate(fake_data, 2):
            for col, val in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col, value=val)

        # Embed beacon as a hyperlink in a hidden cell
        ws.cell(row=100, column=1, value="=")
        ws.cell(row=100, column=1).hyperlink = beacon_url

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    except ImportError:
        logger.warning("openpyxl not installed — generating HTML bait instead")
        return _generate_html_bait(token_id, beacon_url)


def _generate_pdf_bait(token_id: str, beacon_url: str) -> bytes:
    """
    Generate a PDF with an embedded URL action beacon.
    Falls back to HTML if reportlab not available.
    """
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.pdfgen import canvas as pdf_canvas

        buf = BytesIO()
        c = pdf_canvas.Canvas(buf, pagesize=letter)

        c.setFont("Helvetica-Bold", 18)
        c.drawString(72, 720, "Network Infrastructure — CONFIDENTIAL")

        c.setFont("Helvetica", 11)
        c.drawString(72, 690, "Internal Server Inventory — Q1 2024")
        c.drawString(72, 670, "Classification: RESTRICTED — Do not distribute")

        # Fake server table
        c.setFont("Helvetica-Bold", 10)
        y = 630
        c.drawString(72, y, "Hostname")
        c.drawString(200, y, "IP Address")
        c.drawString(320, y, "Service")
        c.drawString(420, y, "Credentials")

        c.setFont("Helvetica", 9)
        servers = [
            ("prod-db-01", "10.0.1.50", "PostgreSQL", "dbadmin:Pr0d#2024"),
            ("prod-web-01", "10.0.1.10", "Nginx", "deploy:D3pl0y!Key"),
            ("prod-cache", "10.0.1.30", "Redis", "AUTH: R3d!s$ecret"),
            ("backup-nas", "10.0.2.100", "NFS/SFTP", "backup:B@ckup2024"),
            ("vpn-gateway", "10.0.0.1", "OpenVPN", "admin:VpN#Master"),
        ]
        for i, (host, ip, svc, cred) in enumerate(servers):
            y = 610 - (i * 20)
            c.drawString(72, y, host)
            c.drawString(200, y, ip)
            c.drawString(320, y, svc)
            c.drawString(420, y, cred)

        # Embed beacon as a URI action
        from reportlab.lib.units import inch
        c.linkURL(beacon_url, (0, 0, 0.1 * inch, 0.1 * inch))

        c.save()
        buf.seek(0)
        return buf.read()

    except ImportError:
        logger.warning("reportlab not installed — generating HTML bait instead")
        return _generate_html_bait(token_id, beacon_url)
